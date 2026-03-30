import os
import tkinter as tk
from tkinter import messagebox

from openpyxl import Workbook, load_workbook  # aggiunto load_workbook


class ModuloBase:
    def __init__(self):
        self.workbook = Workbook()
        self.root = tk.Tk()
        self.root.title("Finestra Base")
        self.root.geometry("800x350")
        self.root.configure(bg="#2c3e50")

        self.campi = {}
        self.crea_interfaccia()

# creazione del placeholder

    def set_placeholder(self, entry, testo):
        entry.insert(0, testo)          # mostra il placeholder subito
        entry.config(fg="gray")

        def on_focus_in(event):
            if entry.get() == testo:
                entry.delete(0, tk.END)
                entry.config(fg="black")

        def on_focus_out(event):
            if entry.get() == "":
                entry.insert(0, testo)
                entry.config(fg="gray")

        entry.bind("<FocusIn>", on_focus_in)
        entry.bind("<FocusOut>", on_focus_out)

 # creazione dei campi di input
 
    def crea_campo(self, testo, placeholder, riga, colonna):
        label = tk.Label(self.root, text=testo, bg="#e5e7eb", font=("Arial", 11))
        label.grid(row=riga, column=colonna, sticky="w", padx=20, pady=(60, 2))

        entry = tk.Entry(self.root, font=("Arial", 12), bg="#cbd5e1", relief="flat", width=25)
        entry.grid(row=riga, column=colonna + 1, padx=10, pady=(60, 10), ipady=8)
        self.set_placeholder(entry, placeholder)
        self.campi[testo] = entry

    def crea_interfaccia(self):
        self.crea_campo("Nome", "Inserisci il tuo nome", 1, 0)
        self.crea_campo("Cognome", "Inserisci il tuo cognome", 1, 2)
        self.crea_campo("Email", "Inserisci la tua email", 3, 0)
        self.crea_campo("Data di nascita", "Inserisci la tua data di nascita", 3, 2)

        btn = tk.Button(
            self.root, text="Salva", command=self.salva_excel,
            bg="#3498db", fg="white", font=("Arial", 12), relief="flat"
        )
        btn.grid(row=4, column=0, columnspan=2, pady=20)

# leggi il testo


    def leggi_dati(self):
        valori = []
        for nome, entry in self.campi.items():   # .items() con parentesi
            testo = entry.get()
            if entry.cget("fg") == "gray":       # placeholder attivo
                testo = ""
            valori.append(testo)
        return valori

    def salva_excel(self):
        dati = self.leggi_dati()
        file = "dati.xlsx"

        if os.path.exists(file):
            wb = load_workbook(file)   # funzione importata, non metodo
            ws = wb.active             # variabile locale wb
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(list(self.campi.keys()))

        ws.append(dati)
        wb.save(file)
        messagebox.showinfo("Salvataggio completato", "I dati sono stati salvati nel file Excel.")

    def avvia(self):
        self.root.mainloop()


app = ModuloBase()
app.avvia()