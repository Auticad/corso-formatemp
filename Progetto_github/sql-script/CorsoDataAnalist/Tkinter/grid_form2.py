import os
import tkinter as tk
from tkinter import messagebox

from openpyxl import Workbook, load_workbook

# Nome del file Excel
FILE_EXCEL = "dati_studenti.xlsx"

def inizializza_excel():
    """Crea il file Excel con le intestazioni se non esiste."""
    if not os.path.exists(FILE_EXCEL):
        try:
            wb = Workbook()   # Crea un nuovo file Excel in memoria
            ws = wb.active    # Ottieni il foglio attivo
            ws.title = "Dati" # Assegna un titolo al foglio
            # Intestazioni delle colonne
            ws.append(["Nome", "Cognome", "Email", "Data di Nascita"])  # Aggiungi intestazioni delle colonne
            wb.save(FILE_EXCEL) # Salva il file
        except Exception as e:
            messagebox.showerror("Errore Critico", f"Impossibile creare il file Excel: {e}")

def set_placeholder(entry, testo):
    """Gestisce il testo placeholder nei campi di input."""
    entry.insert(0, testo)
    entry.config(fg="gray")

    def on_focus_in(event):
        if entry.get() == testo:
            entry.delete(0, tk.END)
            entry.config(fg="black")

    def on_focus_out(event):
        if entry.get() == "":
            entry.insert(0, testo)
            entry.config(fg="gray")

    entry.bind("<FocusIn>", on_focus_in)
    entry.bind("<FocusOut>", on_focus_out)

def crea_campo(testo, placeholder, riga, colonna):
    """Crea una label e un campo entry associato."""
    label = tk.Label(root, text=testo, bg="#e5e7eb", font=("Arial", 11))
    label.grid(row=riga, column=colonna, sticky="w", padx=20, pady=(60, 2))

    entry = tk.Entry(
        root,
        font=("Arial", 12),
        bg="#cbd5e1",
        relief="flat",
        width=25
    )
    entry.grid(row=riga, column=colonna+1, padx=10, pady=(60, 10), ipady=8)
    set_placeholder(entry, placeholder)
    return entry

def pulisci_campi():
    """Resetta tutti i campi al loro stato iniziale (placeholder)."""
    campi_config = [
        (nome, "Inserisci il Nome"),
        (cognome, "Inserisci il Cognome"),
        (email, "Inserisci la mail"),
        (data, "Inserisci la data di nascita")
    ]
    for entry, placeholder in campi_config:
        entry.delete(0, tk.END)
        entry.insert(0, placeholder)
        entry.config(fg="gray")

def salva_dati():
    """Recupera i dati, li valida e li salva nel file Excel."""
    campi = [nome, cognome, email, data]
    placeholders = ["Inserisci il Nome", "Inserisci il Cognome", "Inserisci la mail", "Inserisci la data di nascita"]
    valori = []

    # Estrazione valori puliti
    for i, campo in enumerate(campi):
        testo = campo.get()
        # Se il colore è grigio o il testo corrisponde al placeholder, consideralo vuoto
        if campo.cget("fg") == "gray" or testo == placeholders[i]:
            testo = ""
        valori.append(testo)

    # Validazione base (almeno Nome e Cognome richiesti)
    if not valori[0] or not valori[1]:
        messagebox.showwarning("Attenzione", "Nome e Cognome sono campi obbligatori.")
        return

    try:
        # Carica il workbook esistente
        wb = load_workbook(FILE_EXCEL)   # Carica il file esistente
        ws = wb.active
        
        # Aggiunge una nuova riga
        ws.append(valori)                # Aggiunge la nuova riga
        
        # Salva e chiude
        wb.save(FILE_EXCEL)              # Salva il file
        
        # Feedback positivo
        messagebox.showinfo("Successo", "Dati salvati correttamente nel file Excel!")  # Mostra un messaggio di successo
        
        # Pulisce il modulo per il prossimo inserimento
        pulisci_campi()

    except PermissionError:
        messagebox.showerror("Errore", "Impossibile salvare: il file Excel è aperto in un altro programma. Chiudilo e riprova.")
    except Exception as e:
        messagebox.showerror("Errore", f"Si è verificato un errore durante il salvataggio:\n{e}")

# --- Configurazione Interfaccia Grafica ---

root = tk.Tk()
root.title("Modulo Scolastico")
root.configure(bg="#87a2d8")
root.geometry("800x400") # Leggermente aumentato per spazio

# Inizializza il file Excel all'avvio
inizializza_excel()


# Campi
nome    = crea_campo("Nome",            "Inserisci il Nome",           1, 0)
cognome = crea_campo("Cognome",         "Inserisci il Cognome",        1, 2)
email   = crea_campo("Email",           "Inserisci la mail",           3, 0)
data    = crea_campo("Data di Nascita", "Inserisci la data di nascita",3, 2)

# Bottone Salva
btn = tk.Button(
    root,
    text="Salva Dati",
    bg="#94a3b8",
    activebackground="#64748b",
    fg="white",
    font=("Arial", 11, "bold"),
    command=salva_dati,
    width=20,
    height=2
)
btn.grid(row=4, column=1, columnspan=2, padx=20, pady=20, sticky="w")

btn_chiudi = tk.Button(
    root,
    text="Chiudi",
    bg="#ef4444",
    activebackground="#dc2626",
    fg="white",
    font=("Arial", 11, "bold"),
    command=root.destroy,  # ← Chiude direttamente
    width=20,
    height=2
)
btn_chiudi.grid(row=4, column=3, columnspan=2, padx=20, pady=10, sticky="w")

# Avvio loop principale
root.mainloop()